"""Excel 导出模块"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_to_excel(all_results: list, output_path: str):
    """将所有文件的处理结果导出为格式化 Excel

    Args:
        all_results: list of dict，每个 dict 包含文件名变量、元数据、PWM 台阶结果
        output_path: 输出 Excel 文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "力效测试汇总"

    # 定义列顺序
    columns = [
        "文件名",
        "日期", "桨叶转向", "电机型号", "桨叶材质", "批次", "桨叶序号", "运行序号",
        "涵道型号", "涵道唇口半径", "涵道延伸段长度", "涵道延伸段角度", "涵道间隙",
        "环境温度(℃)", "环境湿度(%RH)", "空气密度(kg/m³)",
        "PWM-μs", "油门-%", "拉力-g", "扭矩-N·m", "电功率-W", "拉力力效-g/W",
        "样本数", "剔除数",
    ]

    # 写入表头
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 写入数据
    row_idx = 2
    for file_result in all_results:
        file_info = file_result["file_info"]
        metadata = file_result["metadata"]
        filename = file_result["filename"]
        pwm_result = file_result.get("pwm_result")

        if pwm_result is None:
            continue

        values = [
            filename,
            file_info.get("日期"), file_info.get("桨叶转向"), file_info.get("电机型号"),
            file_info.get("桨叶材质"), file_info.get("批次"), file_info.get("桨叶序号"),
            file_info.get("运行序号"),
            file_info.get("涵道型号"), file_info.get("涵道唇口半径"),
            file_info.get("涵道延伸段长度"), file_info.get("涵道延伸段角度"),
            file_info.get("涵道间隙"),
            _to_float(metadata.get("环境温度")),
            _to_float(metadata.get("环境湿度")),
            _to_float(metadata.get("空气密度")),
            pwm_result["PWM-μs"],
            pwm_result["油门-%"],
            pwm_result["拉力-g"],
            pwm_result["扭矩-N·m"],
            pwm_result["电功率-W"],
            pwm_result["拉力力效-g/W"],
            pwm_result["样本数"],
            pwm_result["剔除数"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动调整列宽（近似）
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name) * 2  # 中文字符宽
        for row in range(2, min(row_idx, 50)):  # 采样前50行
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)) * 2)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    wb.save(output_path)


def _to_float(value) -> float:
    """安全转换为 float"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None
